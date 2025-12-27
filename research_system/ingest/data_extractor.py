"""
Data file metadata extraction.

Parses data files (.csv, .xlsx, .parquet, .json) to extract:
- Structural info: columns, row count, date range
- Uses LLM to generate: name, type classification, description, tags
"""

import csv
import json
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
from dataclasses import dataclass, field
from datetime import datetime

from research_system.llm.client import LLMClient


# LLM prompts for data file classification
DATA_EXTRACTION_SYSTEM_PROMPT = """You are a data analyst for a trading research system. Your task is to analyze data file previews and classify them.

Classify the data type as one of:
- price_data: OHLCV, price series, returns, quotes
- breadth_data: Market breadth indicators (advance/decline, new highs/lows, McClellan)
- fundamental: Earnings, revenue, balance sheet, financial ratios
- alternative: Sentiment, satellite, social media, unusual data
- reference: Static reference data (ticker mappings, sector classifications, calendars)
- event_data: Corporate events, economic releases, fed meetings

Be specific with naming and tagging. Extract meaning from column names and sample values."""

DATA_EXTRACTION_USER_PROMPT = """Analyze this data file preview and extract metadata.

Filename: {filename}
Format: {format}
Columns: {columns}
Row count: {row_count}
Date range: {date_range}

Sample rows (first 5):
{sample_rows}

Extract the following in JSON format:
{{
  "name": "Human-readable name (max 50 chars)",
  "type": "price_data|breadth_data|fundamental|alternative|reference|event_data",
  "description": "One-line description of what this data contains (max 200 chars)",
  "tags": ["categorization", "tags"]
}}

Respond with ONLY the JSON object, no other text."""


@dataclass
class DataFileInfo:
    """Parsed information from a data file."""
    columns: List[str]
    row_count: int
    sample_rows: List[Dict[str, Any]] = field(default_factory=list)
    date_column: Optional[str] = None
    date_range: Optional[Tuple[str, str]] = None  # (start, end)
    file_format: str = "csv"


@dataclass
class DataExtractionResult:
    """Result of data file extraction."""
    success: bool
    file_info: Optional[DataFileInfo] = None
    metadata: Optional[Dict[str, Any]] = None
    raw_response: Optional[str] = None
    error: Optional[str] = None


class DataFileExtractor:
    """
    Extracts metadata from data files.

    Supports: .csv, .xls, .xlsx, .parquet, .json (array-of-objects)
    """

    SUPPORTED_EXTENSIONS = {".csv", ".xls", ".xlsx", ".parquet", ".json"}

    # Patterns for detecting date columns
    DATE_COLUMN_PATTERNS = ["date", "time", "timestamp", "dt", "datetime", "day", "month", "year"]

    # Maximum rows to read for sampling
    MAX_SAMPLE_ROWS = 5
    MAX_ROWS_FOR_RANGE = 100000  # Read up to this many rows to find date range

    def __init__(self, llm_client: Optional[LLMClient] = None):
        """
        Initialize the data file extractor.

        Args:
            llm_client: LLM client for metadata generation. If None, returns file info only.
        """
        self.llm_client = llm_client

    def extract(self, file_path: Path) -> DataExtractionResult:
        """
        Extract metadata from a data file.

        Args:
            file_path: Path to the data file

        Returns:
            DataExtractionResult with file info and LLM-generated metadata
        """
        suffix = file_path.suffix.lower()

        if suffix not in self.SUPPORTED_EXTENSIONS:
            return DataExtractionResult(
                success=False,
                error=f"Unsupported data file type: {suffix}"
            )

        # Parse file to get structure
        try:
            if suffix == ".csv":
                file_info = self._parse_csv(file_path)
            elif suffix in (".xls", ".xlsx"):
                file_info = self._parse_excel(file_path)
            elif suffix == ".parquet":
                file_info = self._parse_parquet(file_path)
            elif suffix == ".json":
                file_info = self._parse_json(file_path)
            else:
                return DataExtractionResult(
                    success=False,
                    error=f"No parser for: {suffix}"
                )
        except Exception as e:
            return DataExtractionResult(
                success=False,
                error=f"Failed to parse file: {str(e)}"
            )

        if not file_info.columns:
            return DataExtractionResult(
                success=False,
                error="No columns found in data file"
            )

        if file_info.row_count == 0:
            return DataExtractionResult(
                success=False,
                error="Data file is empty (no rows)"
            )

        # Detect date column and extract range
        file_info.date_column = self._detect_date_column(file_info.columns, file_info.sample_rows)

        # Extract with LLM if available
        if self.llm_client:
            metadata = self._extract_with_llm(file_info, file_path.name)
            if metadata:
                return DataExtractionResult(
                    success=True,
                    file_info=file_info,
                    metadata=metadata
                )

        # Return file info without LLM metadata
        return DataExtractionResult(
            success=True,
            file_info=file_info,
            metadata={
                "name": file_path.stem.replace("_", " ").title(),
                "type": "unknown",
                "description": f"Data file with {file_info.row_count} rows and {len(file_info.columns)} columns",
                "tags": []
            }
        )

    def _parse_csv(self, file_path: Path) -> DataFileInfo:
        """Parse a CSV file."""
        with open(file_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
            # Try to detect delimiter
            sample = f.read(8192)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel  # Default to comma-separated

            reader = csv.DictReader(f, dialect=dialect)
            columns = reader.fieldnames or []

            sample_rows = []
            row_count = 0
            first_date = None
            last_date = None
            date_col = self._detect_date_column(columns, [])

            for row in reader:
                row_count += 1
                if len(sample_rows) < self.MAX_SAMPLE_ROWS:
                    sample_rows.append(dict(row))

                # Track date range
                if date_col and date_col in row and row_count <= self.MAX_ROWS_FOR_RANGE:
                    date_val = row[date_col]
                    if date_val:
                        if first_date is None:
                            first_date = date_val
                        last_date = date_val

        date_range = None
        if first_date and last_date:
            date_range = (str(first_date), str(last_date))

        return DataFileInfo(
            columns=list(columns),
            row_count=row_count,
            sample_rows=sample_rows,
            date_column=date_col,
            date_range=date_range,
            file_format="csv"
        )

    def _parse_excel(self, file_path: Path) -> DataFileInfo:
        """Parse an Excel file (.xls, .xlsx)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl required for Excel files: pip install openpyxl")

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return DataFileInfo(columns=[], row_count=0, file_format="excel")

        # First row is headers
        columns = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = rows[1:]

        # Convert to list of dicts
        sample_rows = []
        for row in data_rows[:self.MAX_SAMPLE_ROWS]:
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i] if i < len(row) else None
                row_dict[col] = val
            sample_rows.append(row_dict)

        # Detect date column and range
        date_col = self._detect_date_column(columns, sample_rows)
        date_range = None

        if date_col and data_rows:
            col_idx = columns.index(date_col)
            first_val = data_rows[0][col_idx] if col_idx < len(data_rows[0]) else None
            last_val = data_rows[-1][col_idx] if col_idx < len(data_rows[-1]) else None
            if first_val and last_val:
                date_range = (str(first_val), str(last_val))

        wb.close()

        return DataFileInfo(
            columns=columns,
            row_count=len(data_rows),
            sample_rows=sample_rows,
            date_column=date_col,
            date_range=date_range,
            file_format="excel"
        )

    def _parse_parquet(self, file_path: Path) -> DataFileInfo:
        """Parse a Parquet file."""
        try:
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError("pyarrow required for Parquet files: pip install pyarrow")

        table = pq.read_table(file_path)
        columns = table.column_names
        row_count = table.num_rows

        # Get sample rows
        sample_table = table.slice(0, min(self.MAX_SAMPLE_ROWS, row_count))
        sample_rows = sample_table.to_pylist()

        # Detect date column and range
        date_col = self._detect_date_column(columns, sample_rows)
        date_range = None

        if date_col and row_count > 0:
            # Read first and last values for date range
            col_data = table.column(date_col)
            first_val = col_data[0].as_py() if len(col_data) > 0 else None
            last_val = col_data[-1].as_py() if len(col_data) > 0 else None
            if first_val and last_val:
                date_range = (str(first_val), str(last_val))

        return DataFileInfo(
            columns=list(columns),
            row_count=row_count,
            sample_rows=sample_rows,
            date_column=date_col,
            date_range=date_range,
            file_format="parquet"
        )

    def _parse_json(self, file_path: Path) -> DataFileInfo:
        """
        Parse a JSON data file.

        Expects array of objects: [{"col1": val1, ...}, ...]
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not isinstance(data, list):
            raise ValueError("JSON data file must be an array of objects")

        if not data:
            return DataFileInfo(columns=[], row_count=0, file_format="json")

        if not isinstance(data[0], dict):
            raise ValueError("JSON data file must contain objects (dicts)")

        # Get columns from first object
        columns = list(data[0].keys())
        row_count = len(data)

        sample_rows = data[:self.MAX_SAMPLE_ROWS]

        # Detect date column and range
        date_col = self._detect_date_column(columns, sample_rows)
        date_range = None

        if date_col and row_count > 0:
            first_val = data[0].get(date_col)
            last_val = data[-1].get(date_col)
            if first_val and last_val:
                date_range = (str(first_val), str(last_val))

        return DataFileInfo(
            columns=columns,
            row_count=row_count,
            sample_rows=sample_rows,
            date_column=date_col,
            date_range=date_range,
            file_format="json"
        )

    def _detect_date_column(self, columns: List[str], sample_rows: List[Dict]) -> Optional[str]:
        """
        Detect which column contains dates.

        Uses column name patterns and optionally validates with sample data.
        """
        columns_lower = {col: col.lower() for col in columns}

        # Check for exact matches first
        for col, col_lower in columns_lower.items():
            if col_lower in self.DATE_COLUMN_PATTERNS:
                return col

        # Check for partial matches
        for col, col_lower in columns_lower.items():
            for pattern in self.DATE_COLUMN_PATTERNS:
                if pattern in col_lower:
                    return col

        return None

    def _extract_with_llm(self, file_info: DataFileInfo, filename: str) -> Optional[Dict[str, Any]]:
        """Use LLM to generate metadata from file preview."""
        if self.llm_client is None:
            return None

        # Format sample rows for display
        sample_str = ""
        for i, row in enumerate(file_info.sample_rows[:5], 1):
            row_items = [f"{k}: {v}" for k, v in list(row.items())[:8]]  # Limit columns shown
            sample_str += f"  {i}. {{{', '.join(row_items)}}}\n"

        date_range_str = "Unknown"
        if file_info.date_range:
            date_range_str = f"{file_info.date_range[0]} to {file_info.date_range[1]}"

        prompt = DATA_EXTRACTION_USER_PROMPT.format(
            filename=filename,
            format=file_info.file_format,
            columns=", ".join(file_info.columns),
            row_count=file_info.row_count,
            date_range=date_range_str,
            sample_rows=sample_str
        )

        response = self.llm_client.generate_haiku(
            user=prompt,
            system=DATA_EXTRACTION_SYSTEM_PROMPT,
            max_tokens=512
        )

        if response.offline:
            return None

        metadata = self.llm_client.extract_json(response)

        if metadata is None:
            return None

        # Validate required fields
        if not metadata.get("name"):
            metadata["name"] = filename.rsplit(".", 1)[0].replace("_", " ").title()

        valid_types = ["price_data", "breadth_data", "fundamental", "alternative", "reference", "event_data"]
        if metadata.get("type") not in valid_types:
            metadata["type"] = "unknown"

        if not metadata.get("tags"):
            metadata["tags"] = []

        return metadata


def is_data_json(file_path: Path) -> bool:
    """
    Check if a JSON file contains tabular data (array of uniform objects).

    Used to distinguish between:
    - Data files: [{"col1": v1, "col2": v2}, {...}, ...]
    - Document files: {"type": "strategy", "name": "..."}
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            # Read just enough to determine structure
            content = f.read(65536)  # 64KB should be enough

        data = json.loads(content)

        # Data files are arrays of objects
        if not isinstance(data, list):
            return False

        if len(data) < 2:
            return False

        # Check if first two elements are dicts with same keys
        if not isinstance(data[0], dict) or not isinstance(data[1], dict):
            return False

        return set(data[0].keys()) == set(data[1].keys())

    except (json.JSONDecodeError, IOError):
        return False
